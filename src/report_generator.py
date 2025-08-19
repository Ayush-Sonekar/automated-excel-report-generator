"""
Excel Report Generation Module
Creates professional Excel reports with formatting, charts, and multiple sheets.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from datetime import datetime
import os
from chart_generator import ChartGenerator
from config import Config

class ReportGenerator:
    """Generates professional Excel reports with charts and formatting."""
    
    def __init__(self, verbose=False):
        self.verbose = verbose
        self.config = Config()
        self.chart_generator = ChartGenerator()
        
        # Define styling
        self.header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')
        self.data_font = Font(name='Arial', size=10)
        self.title_font = Font(name='Arial', size=16, bold=True, color='2E86AB')
        
        # Border styles
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def create_workbook_structure(self):
        """
        Create the basic workbook structure with multiple sheets.
        
        Returns:
            Workbook: Configured workbook object
        """
        wb = Workbook()
        
        # Remove default sheet and create new ones
        if wb.active:
            wb.remove(wb.active)
        
        # Create sheets
        wb.create_sheet("Executive Summary", 0)
        wb.create_sheet("Raw Data", 1)
        wb.create_sheet("Monthly Analysis", 2)
        wb.create_sheet("Product Analysis", 3)
        wb.create_sheet("Regional Analysis", 4)
        wb.create_sheet("Charts", 5)
        
        return wb
    
    def format_header_row(self, ws, row_num, start_col=1, end_col=None):
        """
        Format a row as a header with styling.
        
        Args:
            ws: Worksheet object
            row_num (int): Row number to format
            start_col (int): Starting column
            end_col (int): Ending column (auto-detect if None)
        """
        if end_col is None:
            # Auto-detect end column based on content
            end_col = ws.max_column
        
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.thin_border
    
    def format_data_range(self, ws, start_row, start_col, end_row, end_col):
        """
        Format a data range with consistent styling.
        
        Args:
            ws: Worksheet object
            start_row, start_col, end_row, end_col: Range boundaries
        """
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = self.data_font
                cell.border = self.thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def add_title(self, ws, title, row=1, col=1):
        """
        Add a formatted title to a worksheet.
        
        Args:
            ws: Worksheet object
            title (str): Title text
            row (int): Row position
            col (int): Column position
        """
        title_cell = ws.cell(row=row, column=col)
        title_cell.value = title
        title_cell.font = self.title_font
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Add timestamp
        timestamp_cell = ws.cell(row=row + 1, column=col)
        timestamp_cell.value = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        timestamp_cell.font = Font(name='Arial', size=10, italic=True)
    
    def write_dataframe_to_sheet(self, ws, df, start_row=4, title=None):
        """
        Write a dataframe to a worksheet with proper formatting.
        
        Args:
            ws: Worksheet object
            df (pd.DataFrame): Data to write
            start_row (int): Starting row
            title (str): Optional title
            
        Returns:
            int: Last row used
        """
        current_row = start_row
        
        # Add title if provided
        if title:
            self.add_title(ws, title, current_row)
            current_row += 3
        
        if df.empty:
            ws.cell(row=current_row, column=1).value = "No data available"
            return current_row + 1
        
        # Write dataframe
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=current_row + r_idx, column=c_idx)
                cell.value = value
        
        # Format header row
        if len(df) > 0:
            self.format_header_row(ws, current_row, 1, len(df.columns))
            
            # Format data rows
            if len(df) > 0:
                data_start_row = current_row + 1
                data_end_row = current_row + len(df)
                self.format_data_range(ws, data_start_row, 1, data_end_row, len(df.columns))
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column].width = adjusted_width
        
        return current_row + len(df) + 2
    
    def create_executive_summary(self, ws, processed_data):
        """
        Create executive summary sheet with key insights.
        
        Args:
            ws: Worksheet object
            processed_data (dict): Processed data dictionary
        """
        # Add domain-specific title
        domain_info = processed_data.get('domain_info', {})
        domain = domain_info.get('domain', 'business').replace('_', ' ').title()
        confidence = domain_info.get('confidence', 0)
        
        title = f"Executive Summary - {domain} Data Analysis"
        if confidence > 70:
            title += f" (High Confidence: {confidence:.0f}%)"
        
        self.add_title(ws, title)
        
        current_row = 4
        summary_stats = processed_data['aggregations'].get('summary', {})
        
        # Key metrics section
        metrics = [
            ("Total Records Processed", f"{summary_stats.get('total_records', 'N/A'):,}"),
            ("Data Period", summary_stats.get('date_range', 'N/A')),
            ("Total Sales", f"${summary_stats.get('total_sales', 0):,.2f}" if summary_stats.get('total_sales') else 'N/A'),
            ("Average Sale Value", f"${summary_stats.get('average_sale', 0):.2f}" if summary_stats.get('average_sale') else 'N/A'),
        ]
        
        # Write metrics
        for i, (metric, value) in enumerate(metrics):
            ws.cell(row=current_row + i, column=1).value = metric
            ws.cell(row=current_row + i, column=1).font = Font(name='Arial', size=11, bold=True)
            ws.cell(row=current_row + i, column=3).value = value
            ws.cell(row=current_row + i, column=3).font = Font(name='Arial', size=11)
        
        current_row += len(metrics) + 2
        
        # Top insights section
        ws.cell(row=current_row, column=1).value = "Key Insights"
        ws.cell(row=current_row, column=1).font = self.title_font
        current_row += 2
        
        # Use intelligent insights if available
        insights = processed_data.get('insights', [])
        
        if not insights:
            # Fallback to basic insights if intelligence module not used
            if 'product' in processed_data['aggregations'] and not processed_data['aggregations']['product'].empty:
                top_product = processed_data['aggregations']['product'].iloc[0]
                insights.append(f"• Top performing product: {top_product.iloc[0]} (${top_product.iloc[1]:,.2f} in sales)")
            
            if 'monthly' in processed_data['aggregations'] and not processed_data['aggregations']['monthly'].empty:
                monthly_data = processed_data['aggregations']['monthly']
                best_month = monthly_data.loc[monthly_data.iloc[:, 1].idxmax()]
                insights.append(f"• Best performing month: {best_month.iloc[0]} (${best_month.iloc[1]:,.2f} in sales)")
            
            if 'regional' in processed_data['aggregations'] and not processed_data['aggregations']['regional'].empty:
                top_region = processed_data['aggregations']['regional'].iloc[0]
                insights.append(f"• Top performing region: {top_region.iloc[0]} (${top_region.iloc[1]:,.2f} in sales)")
            
            if not insights:
                insights.append("• Analysis complete - detailed breakdowns available in individual sheets")
        
        # Format insights with bullet points
        formatted_insights = []
        for insight in insights[:8]:  # Limit to 8 insights for space
            if not insight.startswith('•'):
                insight = f"• {insight}"
            formatted_insights.append(insight)
        
        # Write insights
        for i, insight in enumerate(formatted_insights):
            ws.cell(row=current_row + i, column=1).value = insight
            ws.cell(row=current_row + i, column=1).font = Font(name='Arial', size=10)
    
    def add_charts_to_sheet(self, ws, charts):
        """
        Add generated charts to a worksheet.
        
        Args:
            ws: Worksheet object
            charts (dict): Dictionary of chart buffers
        """
        self.add_title(ws, "Data Visualizations")
        
        row_position = 4
        temp_files_to_cleanup = []
        
        for chart_name, chart_buffer in charts.items():
            if chart_buffer:
                try:
                    # Save chart buffer to temporary file
                    temp_file = f"temp_chart_{chart_name}.png"
                    temp_files_to_cleanup.append(temp_file)
                    
                    with open(temp_file, 'wb') as f:
                        chart_buffer.seek(0)
                        f.write(chart_buffer.read())
                    
                    # Insert image into worksheet (don't delete temp file yet)
                    img = Image(temp_file)
                    
                    ws.add_image(img, f'A{row_position}')
                    row_position += 25  # Leave space for next chart
                    
                except Exception as e:
                    if self.verbose:
                        print(f"Error adding chart {chart_name}: {e}")
        
        # Store temp files for cleanup after workbook is saved
        self._temp_files_to_cleanup = temp_files_to_cleanup
    
    def generate_report(self, processed_data, output_path):
        """
        Generate the complete Excel report.
        
        Args:
            processed_data (dict): Processed data dictionary
            output_path (str): Output file path
            
        Returns:
            bool: Success status
        """
        try:
            if self.verbose:
                print("Creating workbook structure...")
            
            # Create workbook
            wb = self.create_workbook_structure()
            
            # Generate charts
            if self.verbose:
                print("Generating charts...")
            charts = self.chart_generator.generate_all_charts(processed_data['aggregations'])
            
            # Executive Summary Sheet
            if self.verbose:
                print("Creating executive summary...")
            executive_ws = wb["Executive Summary"]
            self.create_executive_summary(executive_ws, processed_data)
            
            # Raw Data Sheet
            if self.verbose:
                print("Writing raw data...")
            raw_ws = wb["Raw Data"]
            self.write_dataframe_to_sheet(raw_ws, processed_data['raw_data'], title="Raw Business Data")
            
            # Monthly Analysis Sheet
            if 'monthly' in processed_data['aggregations']:
                if self.verbose:
                    print("Creating monthly analysis...")
                monthly_ws = wb["Monthly Analysis"]
                self.write_dataframe_to_sheet(monthly_ws, processed_data['aggregations']['monthly'], 
                                            title="Monthly Sales Analysis")
            
            # Product Analysis Sheet
            if 'product' in processed_data['aggregations']:
                if self.verbose:
                    print("Creating product analysis...")
                product_ws = wb["Product Analysis"]
                self.write_dataframe_to_sheet(product_ws, processed_data['aggregations']['product'], 
                                            title="Product Performance Analysis")
            
            # Regional Analysis Sheet
            if 'regional' in processed_data['aggregations']:
                if self.verbose:
                    print("Creating regional analysis...")
                regional_ws = wb["Regional Analysis"]
                self.write_dataframe_to_sheet(regional_ws, processed_data['aggregations']['regional'], 
                                             title="Regional Sales Analysis")
            
            # Charts Sheet
            if charts:
                if self.verbose:
                    print("Adding charts...")
                charts_ws = wb["Charts"]
                self.add_charts_to_sheet(charts_ws, charts)
            
            # Save workbook
            if self.verbose:
                print(f"Saving report to {output_path}...")
            wb.save(output_path)
            
            # Clean up temporary files after workbook is saved
            if hasattr(self, '_temp_files_to_cleanup'):
                for temp_file in self._temp_files_to_cleanup:
                    try:
                        if os.path.exists(temp_file):
                            os.remove(temp_file)
                    except Exception as e:
                        if self.verbose:
                            print(f"Warning: Could not remove temporary file {temp_file}: {e}")
            
            return True
            
        except Exception as e:
            print(f"Error generating report: {str(e)}")
            if self.verbose:
                import traceback
                traceback.print_exc()
            return False
